import pandas as pd
import openpyxl
from openpyxl import load_workbook
import xlsxwriter

text_file_path = 'out.txt'
excel_file_path = 'MaxCut.xlsx'

wb = load_workbook(excel_file_path)
ws = wb.active

with open(text_file_path, 'r') as f:
    lines = f.readlines()
    
    for line in lines:
        temp = line.split('\t')
        
        input_file_name = temp[0].split('/')[1].strip()
        v = temp[1].split('=')[1].strip()
        e = temp[2].split('=')[1].strip()
        construction_method = temp[3].split('=')[0].strip().split(' ')[0].strip()
        construction_cut = temp[3].split('=')[1].strip()
        local_cut = temp[5].split('=')[1].strip()
        local_iteration = temp[6].split('=')[1].strip()
        GRASP_iteration = None
        GRASP_cut = None
        
        # if construction method is Semi-Greedy, do the following
        if construction_method[0] == 'S':
            GRASP_iteration = temp[7].split('=')[1].strip()
            GRASP_cut = temp[8].split('=')[1].strip()
        
        print(input_file_name, v, e, construction_method, construction_cut, local_cut, local_iteration, GRASP_iteration, GRASP_cut)
        
        input_number = int(input_file_name.split('.')[0][1:])
        row_number = 6 + input_number
        ws['B' + str(row_number)].value = v
        ws['C' + str(row_number)].value = e
        
        if construction_method == 'Randomized':
            ws['D' + str(row_number)].value = construction_cut
            ws['J' + str(row_number)].value = local_cut
            ws['I' + str(row_number)].value = local_iteration
        elif construction_method == 'Semi-Greedy-1':
            ws['F' + str(row_number)].value = construction_cut
            ws['N' + str(row_number)].value = local_cut
            ws['M' + str(row_number)].value = local_iteration
            ws['T' + str(row_number)].value = GRASP_cut
            ws['S' + str(row_number)].value = GRASP_iteration
        elif construction_method == 'Greedy-1':
            ws['E' + str(row_number)].value = construction_cut
            ws['L' + str(row_number)].value = local_cut
            ws['K' + str(row_number)].value = local_iteration
        elif construction_method == 'Semi-Greedy-2':
            ws['H' + str(row_number)].value = construction_cut
            ws['R' + str(row_number)].value = local_cut
            ws['Q' + str(row_number)].value = local_iteration
            ws['V' + str(row_number)].value = GRASP_cut
            ws['U' + str(row_number)].value = GRASP_iteration
        elif construction_method == 'Greedy-2':
            ws['G' + str(row_number)].value = construction_cut
            ws['P' + str(row_number)].value = local_cut
            ws['O' + str(row_number)].value = local_iteration

wb.save(excel_file_path)